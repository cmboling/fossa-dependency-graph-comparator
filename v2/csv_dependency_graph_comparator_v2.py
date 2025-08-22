#!/usr/bin/env python3
"""
CSV Dependency Graph Comparator - Reconstructed Version

This script properly compares dependency graphs by matching source units via OriginPaths
and correctly identifies actual dependency differences vs build environment variations.

Key improvements:
- Source units matched by OriginPaths, not by type
- Proper dependency extraction (direct + transitive)
- Better temporal analysis
- Reduced false positives
- Cleaner Excel reporting with fewer tabs
"""

import json
import sys
import csv
import re
from collections import defaultdict
from typing import Dict, List, Set, Tuple, Any, Optional
import hashlib
from io import StringIO
from datetime import datetime, timezone
from dataclasses import dataclass
import argparse
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Increase CSV field size limit
csv.field_size_limit(10 * 1024 * 1024)

@dataclass
class ParsedLocator:
    """Represents a parsed locator with org, project, and revision."""
    org_id: str
    project_title: str
    revision: str
    original: str
    
    def get_project_key(self) -> str:
        """Get the project key (org + project) for grouping."""
        return f"{self.org_id}/{self.project_title}"

class SourceUnitSignature:
    """Represents a normalized signature for a single source unit."""
    
    def __init__(self, source_unit: Dict[str, Any]):
        self.origin_paths = sorted([path for path in source_unit.get("OriginPaths", []) if path])
        self.type = source_unit.get("Type", "")
        self.build = source_unit.get("Build", {})
        self.succeeded = self.build.get("Succeeded", False)
        
        # Extract direct imports (what this source unit imports)
        raw_imports = self.build.get("Imports", [])
        self.direct_imports = self._normalize_dependency_list(raw_imports)
        
        # Extract transitive dependencies (dependencies and their imports)
        raw_deps = self.build.get("Dependencies", [])
        self.transitive_deps = self._normalize_transitive_dependencies(raw_deps)
        
    def _normalize_dependency_list(self, deps: List[str]) -> Set[str]:
        """Normalize a list of dependency identifiers."""
        normalized = set()
        for dep in deps:
            if dep and isinstance(dep, str):
                # Only normalize build environment paths, not dependency identifiers
                normalized_dep = self._normalize_if_build_path(dep)
                if normalized_dep:
                    normalized.add(normalized_dep)
        return normalized
        
    def _normalize_transitive_dependencies(self, deps: List[Any]) -> Dict[str, Set[str]]:
        """Extract and normalize transitive dependencies properly."""
        normalized = {}
        for dep in deps:
            if isinstance(dep, dict):
                locator = dep.get("locator", "")
                imports = dep.get("imports", [])
                
                if locator:
                    normalized_locator = self._normalize_if_build_path(locator)
                    if normalized_locator:
                        # Normalize the imports as well
                        normalized_imports = set()
                        for imp in imports:
                            if imp and isinstance(imp, str):
                                normalized_imp = self._normalize_if_build_path(imp)
                                if normalized_imp:
                                    normalized_imports.add(normalized_imp)
                        normalized[normalized_locator] = normalized_imports
        
        return normalized
    
    def _normalize_if_build_path(self, path_or_locator: str) -> str:
        """Only normalize if this looks like a build environment path."""
        if not path_or_locator:
            return path_or_locator
            
        # Only normalize paths that look like temporary build paths
        if '/tmp/tmp' in path_or_locator and '/unpacked/' in path_or_locator:
            # Remove temporary directory prefixes
            normalized = re.sub(r'/tmp/tmp[^/]+/unpacked/[^/]+/', '', path_or_locator)
            normalized = normalized.lstrip('/\\')
            return normalized if normalized else path_or_locator
        
        # For dependency locators and normal paths, don't normalize
        return path_or_locator
    
    def get_all_dependencies(self) -> Set[str]:
        """Get all dependencies (direct + transitive) for this source unit."""
        all_deps = set(self.direct_imports)
        
        # Add transitive dependency locators
        all_deps.update(self.transitive_deps.keys())
        
        # Add imports from transitive dependencies
        for imports in self.transitive_deps.values():
            all_deps.update(imports)
            
        return all_deps
    
    def get_signature(self) -> str:
        """Generate a signature for this source unit based on actual dependencies."""
        # Create deterministic representation
        data = {
            "origin_paths": self.origin_paths,
            "type": self.type,
            "build_succeeded": self.succeeded,
            "direct_imports": sorted(list(self.direct_imports)),
            "transitive_deps": {
                locator: sorted(list(imports))
                for locator, imports in sorted(self.transitive_deps.items())
            }
        }
        
        json_str = json.dumps(data, sort_keys=True, separators=(',', ':'))
        return hashlib.sha256(json_str.encode()).hexdigest()
    
    def is_equivalent_to(self, other: 'SourceUnitSignature') -> bool:
        """Check if this source unit is equivalent to another."""
        return (
            self.origin_paths == other.origin_paths and
            self.type == other.type and
            self.succeeded == other.succeeded and
            self.direct_imports == other.direct_imports and
            self.transitive_deps == other.transitive_deps
        )

class ProjectDependencyGraph:
    """Represents a project's dependency graph organized by OriginPaths."""
    
    def __init__(self, locator: str, created_at: str, project_data: Dict[str, Any], build_id: str = ""):
        self.locator = locator
        self.created_at = created_at
        self.build_id = build_id
        self.name = project_data.get("Name", "")
        
        # Group source units by their origin paths
        self.source_units_by_origin = {}
        for su_data in project_data.get("SourceUnits", []):
            source_unit = SourceUnitSignature(su_data)
            
            # Each source unit may have multiple origin paths
            for origin_path in source_unit.origin_paths:
                if origin_path not in self.source_units_by_origin:
                    self.source_units_by_origin[origin_path] = source_unit
                else:
                    # If multiple source units claim the same origin path, this could be an issue
                    print(f"Warning: Multiple source units for origin path {origin_path} in {locator}", file=sys.stderr)
    
    def get_signature(self) -> str:
        """Generate a signature for the entire project dependency graph."""
        # Create signature based on all origin paths and their dependencies
        origin_signatures = {}
        for origin_path, source_unit in sorted(self.source_units_by_origin.items()):
            origin_signatures[origin_path] = source_unit.get_signature()
        
        combined_data = {
            "origin_signatures": origin_signatures,
            "origin_paths": sorted(self.source_units_by_origin.keys())
        }
        
        json_str = json.dumps(combined_data, sort_keys=True, separators=(',', ':'))
        return hashlib.sha256(json_str.encode()).hexdigest()
    
    def is_equivalent_to(self, other: 'ProjectDependencyGraph') -> bool:
        """Check if this project dependency graph is equivalent to another."""
        # Must have same set of origin paths
        if set(self.source_units_by_origin.keys()) != set(other.source_units_by_origin.keys()):
            return False
        
        # For each origin path, source units must be equivalent
        for origin_path in self.source_units_by_origin:
            if not self.source_units_by_origin[origin_path].is_equivalent_to(
                other.source_units_by_origin[origin_path]
            ):
                return False
        
        return True
    
    def get_all_dependencies_flat(self) -> Set[str]:
        """Get all dependencies from all source units as a flat set."""
        all_deps = set()
        for source_unit in self.source_units_by_origin.values():
            all_deps.update(source_unit.get_all_dependencies())
        return all_deps
    
    def get_all_dependencies_summary(self) -> Dict[str, int]:
        """Get summary statistics about dependencies in this graph."""
        total_direct = 0
        total_transitive_locators = 0
        total_transitive_imports = 0
        
        for source_unit in self.source_units_by_origin.values():
            total_direct += len(source_unit.direct_imports)
            total_transitive_locators += len(source_unit.transitive_deps)
            for imports in source_unit.transitive_deps.values():
                total_transitive_imports += len(imports)
        
        return {
            "direct": total_direct,
            "transitive": total_transitive_locators + total_transitive_imports,
            "total": total_direct + total_transitive_locators + total_transitive_imports
        }
    
    def compare_with(self, other: 'ProjectDependencyGraph') -> Dict[str, Any]:
        """Compare this graph with another and return detailed differences."""
        comparison = {
            "are_equivalent": self.is_equivalent_to(other),
            "origin_paths_self": set(self.source_units_by_origin.keys()),
            "origin_paths_other": set(other.source_units_by_origin.keys()),
            "origin_path_differences": {},
            "summary": {
                "total_dependencies_self": 0,
                "total_dependencies_other": 0,
                "common_dependencies": set(),
                "unique_to_self": set(),
                "unique_to_other": set()
            }
        }
        
        # Find origin path differences
        common_origins = comparison["origin_paths_self"] & comparison["origin_paths_other"]
        only_in_self = comparison["origin_paths_self"] - comparison["origin_paths_other"]
        only_in_other = comparison["origin_paths_other"] - comparison["origin_paths_self"]
        
        if only_in_self or only_in_other:
            comparison["origin_path_differences"]["only_in_self"] = list(only_in_self)
            comparison["origin_path_differences"]["only_in_other"] = list(only_in_other)
        
        # Compare dependencies for common origin paths
        all_deps_self = set()
        all_deps_other = set()
        
        for origin_path in common_origins:
            su_self = self.source_units_by_origin[origin_path]
            su_other = other.source_units_by_origin[origin_path]
            
            deps_self = su_self.get_all_dependencies()
            deps_other = su_other.get_all_dependencies()
            
            all_deps_self.update(deps_self)
            all_deps_other.update(deps_other)
            
            if deps_self != deps_other:
                comparison["origin_path_differences"][origin_path] = {
                    "dependencies_self": sorted(list(deps_self)),
                    "dependencies_other": sorted(list(deps_other)),
                    "common": sorted(list(deps_self & deps_other)),
                    "only_in_self": sorted(list(deps_self - deps_other)),
                    "only_in_other": sorted(list(deps_other - deps_self))
                }
        
        # Overall summary
        comparison["summary"]["total_dependencies_self"] = len(all_deps_self)
        comparison["summary"]["total_dependencies_other"] = len(all_deps_other)
        comparison["summary"]["common_dependencies"] = all_deps_self & all_deps_other
        comparison["summary"]["unique_to_self"] = all_deps_self - all_deps_other
        comparison["summary"]["unique_to_other"] = all_deps_other - all_deps_self
        
        return comparison

def parse_locator(locator: str) -> Optional[ParsedLocator]:
    """Parse a locator string to extract org, project, and revision."""
    pattern = r'^custom\+([^/]+)/([^$]+)\$(.+)$'
    match = re.match(pattern, locator)
    
    if not match:
        return None
    
    return ParsedLocator(
        org_id=match.group(1),
        project_title=match.group(2),
        revision=match.group(3),
        original=locator
    )

def parse_timestamp(timestamp_str: str) -> datetime:
    """Parse timestamp string to datetime object."""
    try:
        # Handle various timestamp formats
        if timestamp_str.endswith('Z'):
            return datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
        elif '+' in timestamp_str and ':' in timestamp_str.split('+')[-1]:
            return datetime.fromisoformat(timestamp_str)
        elif timestamp_str.endswith('+00'):
            return datetime.fromisoformat(timestamp_str + ':00')
        else:
            # Assume UTC if no timezone
            dt = datetime.fromisoformat(timestamp_str.replace('Z', ''))
            return dt.replace(tzinfo=timezone.utc)
    except:
        return datetime.now(timezone.utc)

def analyze_project_revisions(grouped_data: Dict[str, Dict[str, List[ProjectDependencyGraph]]]) -> Dict[str, Any]:
    """Analyze projects for alternating dependency graphs within same revisions."""
    results = {
        "high_priority_alternating": [],  # Same revision, different graphs
        "projects_analyzed": 0,
        "revisions_analyzed": 0
    }
    
    for project_key, revisions in grouped_data.items():
        results["projects_analyzed"] += 1
        
        for revision, graphs in revisions.items():
            results["revisions_analyzed"] += 1
            
            if len(graphs) < 2:
                continue  # Need at least 2 analyses to compare
            
            # Group by signature to find alternating patterns
            signature_groups = defaultdict(list)
            for graph in graphs:
                signature = graph.get_signature()
                signature_groups[signature].append(graph)
            
            # If multiple signatures for same revision = alternating!
            if len(signature_groups) > 1:
                # Sort signatures by frequency and time
                sorted_signatures = sorted(
                    signature_groups.items(),
                    key=lambda x: (-len(x[1]), min(g.created_at for g in x[1]))
                )
                
                # Collect ALL build IDs and timestamps from all signature groups
                all_build_ids = []
                all_timestamps = []
                all_comparisons = []
                
                primary_sig, primary_graphs = sorted_signatures[0]
                
                # Create comparisons with each alternate signature
                for alt_sig, alt_graphs in sorted_signatures[1:]:
                    primary_graph = primary_graphs[0]
                    alt_graph = alt_graphs[0]
                    comparison = primary_graph.compare_with(alt_graph)
                    all_comparisons.append({
                        "alt_signature": alt_sig,
                        "alt_count": len(alt_graphs),
                        "alt_build_ids": [g.build_id for g in alt_graphs],
                        "alt_timestamps": [g.created_at for g in alt_graphs],
                        "comparison": comparison,
                        "sample_alt": alt_graph
                    })
                
                # Collect all build IDs and timestamps from all groups
                for sig, sig_graphs in sorted_signatures:
                    all_build_ids.extend([g.build_id for g in sig_graphs])
                    all_timestamps.extend([g.created_at for g in sig_graphs])
                
                # Create ONE entry per project-revision with ALL build information
                results["high_priority_alternating"].append({
                    "project_key": project_key,
                    "revision": revision,
                    "total_analyses": len(graphs),
                    "signature_count": len(signature_groups),
                    "all_build_ids": all_build_ids,
                    "all_timestamps": all_timestamps,
                    "primary_signature": primary_sig,
                    "primary_count": len(primary_graphs),
                    "primary_build_ids": [g.build_id for g in primary_graphs],
                    "primary_timestamps": [g.created_at for g in primary_graphs],
                    "sample_primary": primary_graphs[0],
                    "all_comparisons": all_comparisons,
                    # Keep first alternate for backward compatibility with Excel generation
                    "alt_signature": all_comparisons[0]["alt_signature"],
                    "alt_count": all_comparisons[0]["alt_count"],
                    "alt_build_ids": all_comparisons[0]["alt_build_ids"],
                    "alt_timestamps": all_comparisons[0]["alt_timestamps"],
                    "comparison": all_comparisons[0]["comparison"],
                    "sample_alt": all_comparisons[0]["sample_alt"]
                })
    
    return results

def create_excel_report_v2(analysis: Dict[str, Any], output_path: str):
    """Create improved Excel report with better organization and accurate dependency data."""
    if not EXCEL_AVAILABLE:
        print("Error: openpyxl package not available. Install with: pip install openpyxl", file=sys.stderr)
        return
    
    if not analysis['high_priority_alternating']:
        print("No alternating dependency graphs found. No Excel file created.")
        return
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    graph1_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    graph2_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Create summary sheet
    summary_ws = wb.create_sheet(title="Summary")
    
    # Summary header
    summary_ws.merge_cells('A1:F1')
    summary_ws['A1'] = "Same Revision Alternating Dependency Graphs Analysis"
    summary_ws['A1'].font = Font(bold=True, size=16)
    summary_ws['A1'].alignment = center_align
    
    summary_ws['A3'] = f"Total alternating revisions: {len(analysis['high_priority_alternating'])}"
    summary_ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Summary table
    row = 6
    headers = ["Project", "Revision", "Total Scans", "Unique Signatures", "Build IDs", "Time Range"]
    for col, header in enumerate(headers, 1):
        cell = summary_ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    row += 1
    for item in analysis['high_priority_alternating']:
        # Use the comprehensive all_build_ids that includes ALL builds from ALL signature groups
        all_build_ids = item['all_build_ids']
        all_timestamps = item['all_timestamps']
        time_range = f"{min(all_timestamps)} to {max(all_timestamps)}"
        
        data = [
            item['project_key'],
            item['revision'][:50] + "..." if len(item['revision']) > 50 else item['revision'],
            item['total_analyses'],
            item['signature_count'],
            ", ".join(map(str, all_build_ids)),
            time_range
        ]
        
        for col, value in enumerate(data, 1):
            summary_ws.cell(row=row, column=col, value=str(value))
        row += 1
    
    # Auto-adjust columns
    for column in summary_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create detailed analysis sheet
    details_ws = wb.create_sheet(title="Detailed Analysis")
    
    # Details header
    details_ws.merge_cells('A1:F1')
    details_ws['A1'] = "Detailed Dependency Graph Comparisons"
    details_ws['A1'].font = Font(bold=True, size=14)
    details_ws['A1'].alignment = center_align
    
    row = 3
    for i, item in enumerate(analysis['high_priority_alternating']):
        # Project header
        details_ws.merge_cells(f'A{row}:F{row}')
        details_ws[f'A{row}'] = f"Project {i+1}: {item['project_key']}"
        details_ws[f'A{row}'].font = Font(bold=True, size=12)
        details_ws[f'A{row}'].fill = header_fill
        details_ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        row += 1
        
        # Revision info
        details_ws[f'A{row}'] = f"Revision: {item['revision']}"
        details_ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Get sample graphs for better naming
        sample_graph1 = item['sample_primary']
        sample_graph2 = item['sample_alt']
        
        # Format timestamps for better readability
        time1 = sample_graph1.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(sample_graph1.created_at, 'strftime') else str(sample_graph1.created_at)[:19]
        time2 = sample_graph2.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(sample_graph2.created_at, 'strftime') else str(sample_graph2.created_at)[:19]
        
        # Get actual dependency counts from the sample graphs
        deps1 = sample_graph1.get_all_dependencies_summary()
        deps2 = sample_graph2.get_all_dependencies_summary()
        
        # Graph comparison info with timestamps - keep the preferred naming format
        details_ws[f'A{row}'] = f"Graph from {time1} (Build ID: {sample_graph1.build_id}) - {deps1['total']} total deps ({deps1['direct']} direct, {deps1['transitive']} transitive)"
        details_ws[f'A{row}'].fill = graph1_fill
        
        # Add related builds in adjacent cell
        if item['primary_count'] > 1:
            other_primary_builds = [f"Build {bid} ({ts})" for bid, ts in zip(item['primary_build_ids'], item['primary_timestamps']) if bid != sample_graph1.build_id]
            related_info = f"Related builds with same graph: {', '.join(other_primary_builds)}"
        else:
            related_info = "Only build with this graph signature"
        details_ws[f'B{row}'] = related_info
        details_ws[f'B{row}'].fill = graph1_fill
        details_ws[f'B{row}'].font = Font(italic=True)
        row += 1
        
        details_ws[f'A{row}'] = f"Graph from {time2} (Build ID: {sample_graph2.build_id}) - {deps2['total']} total deps ({deps2['direct']} direct, {deps2['transitive']} transitive)"
        details_ws[f'A{row}'].fill = graph2_fill
        
        # Add related builds in adjacent cell
        if item['alt_count'] > 1:
            other_alt_builds = [f"Build {bid} ({ts})" for bid, ts in zip(item['alt_build_ids'], item['alt_timestamps']) if bid != sample_graph2.build_id]
            related_info = f"Related builds with same graph: {', '.join(other_alt_builds)}"
        else:
            related_info = "Only build with this graph signature"
        details_ws[f'B{row}'] = related_info
        details_ws[f'B{row}'].fill = graph2_fill
        details_ws[f'B{row}'].font = Font(italic=True)
        row += 1
        
        # Get all dependencies from both graphs for accurate comparison
        all_deps1 = sample_graph1.get_all_dependencies_flat()
        all_deps2 = sample_graph2.get_all_dependencies_flat()
        
        common_deps = sorted(list(all_deps1 & all_deps2))
        unique_to_graph1 = sorted(list(all_deps1 - all_deps2))
        unique_to_graph2 = sorted(list(all_deps2 - all_deps1))
        
        # Dependencies comparison headers  
        details_ws[f'A{row}'] = "Common Dependencies"
        details_ws[f'B{row}'] = f"Only in Graph Type 1"
        details_ws[f'C{row}'] = f"Only in Graph Type 2"
        for col in ['A', 'B', 'C']:
            details_ws[f'{col}{row}'].font = header_font
            details_ws[f'{col}{row}'].fill = header_fill
        row += 1
        
        # Dependency lists
        max_deps = max(len(common_deps), len(unique_to_graph1), len(unique_to_graph2), 1)
        
        for i in range(max_deps):
            if i < len(common_deps):
                details_ws[f'A{row + i}'] = common_deps[i]
            if i < len(unique_to_graph1):
                cell = details_ws[f'B{row + i}']
                cell.value = unique_to_graph1[i]
                cell.fill = graph1_fill
            if i < len(unique_to_graph2):
                cell = details_ws[f'C{row + i}']
                cell.value = unique_to_graph2[i]
                cell.fill = graph2_fill
        
        # Add origin path breakdown if different
        row += max_deps + 1
        details_ws[f'A{row}'] = "Origin Path Analysis:"
        details_ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        origins1 = set(sample_graph1.source_units_by_origin.keys())
        origins2 = set(sample_graph2.source_units_by_origin.keys())
        
        if origins1 != origins2:
            details_ws[f'A{row}'] = f"Different origin paths detected!"
            details_ws[f'A{row}'].font = Font(color="FF0000", bold=True)
            row += 1
            details_ws[f'A{row}'] = f"Graph 1 origins: {sorted(list(origins1))}"
            row += 1
            details_ws[f'A{row}'] = f"Graph 2 origins: {sorted(list(origins2))}"
            row += 1
        else:
            details_ws[f'A{row}'] = f"Same origin paths: {sorted(list(origins1))}"
            row += 1
        
        row += 2  # Add spacing between projects
    
    # Auto-adjust columns for details sheet
    for column in details_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        details_ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")

def main():
    parser = argparse.ArgumentParser(description="Analyze CSV dependency graph data for alternating patterns")
    parser.add_argument("csv_file", help="Path to CSV file containing dependency graph data")
    parser.add_argument("--excel", help="Generate Excel report to specified path")
    args = parser.parse_args()
    
    if not os.path.exists(args.csv_file):
        print(f"Error: CSV file '{args.csv_file}' not found", file=sys.stderr)
        return 1
    
    # Parse CSV data
    grouped_data = defaultdict(lambda: defaultdict(list))
    total_rows = 0
    parsed_rows = 0
    
    print("Parsing CSV data...")
    
    with open(args.csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        print(f"Detected header: {header}")
        
        # Determine column indices
        if len(header) == 4:
            build_id_col, locator_col, created_at_col, data_col = 0, 1, 2, 3
            has_build_id = True
        elif len(header) == 3:
            locator_col, created_at_col, data_col = 0, 1, 2
            has_build_id = False
        else:
            print(f"Error: Expected 3 or 4 columns, got {len(header)}", file=sys.stderr)
            return 1
        
        for row_num, row in enumerate(reader, start=2):
            total_rows += 1
            
            try:
                if has_build_id:
                    build_id, locator, created_at, data_json = row
                else:
                    locator, created_at, data_json = row
                    build_id = ""
                
                # Parse data
                parsed_locator = parse_locator(locator)
                if not parsed_locator:
                    continue
                
                parsed_timestamp = parse_timestamp(created_at)
                project_data = json.loads(data_json)
                
                # Create project dependency graph
                graph = ProjectDependencyGraph(locator, parsed_timestamp, project_data, build_id)
                
                # Group by project and revision
                project_key = parsed_locator.get_project_key()
                revision = parsed_locator.revision
                grouped_data[project_key][revision].append(graph)
                
                parsed_rows += 1
                
            except Exception as e:
                print(f"Error processing row {row_num}: {e}", file=sys.stderr)
    
    print(f"Successfully parsed {parsed_rows}/{total_rows} project analyses from CSV data.")
    
    # Analyze for alternating patterns
    analysis = analyze_project_revisions(grouped_data)
    
    # Print results
    print(f"\n🚨 HIGH PRIORITY - Same Revision Alternating Dependency Graphs: {len(analysis['high_priority_alternating'])}")
    print("   These revisions have the SAME code but DIFFERENT dependency analysis results!")
    
    for item in analysis['high_priority_alternating'][:10]:  # Show top 10
        all_build_ids = item['all_build_ids']
        all_timestamps = item['all_timestamps']
        time_range = f"{min(all_timestamps)} to {max(all_timestamps)}"
        
        print(f"   • {item['project_key']} @ {item['revision'][:60]}...")
        print(f"     {item['total_analyses']} scans → {item['signature_count']} different dependency graphs")
        print(f"     Build IDs: {', '.join(map(str, all_build_ids))}")
        print(f"     Time range: {time_range}")
    
    if len(analysis['high_priority_alternating']) > 10:
        print(f"   ... and {len(analysis['high_priority_alternating']) - 10} more alternating revisions")
    
    # Generate Excel report
    if args.excel:
        create_excel_report_v2(analysis, args.excel)

if __name__ == "__main__":
    main()